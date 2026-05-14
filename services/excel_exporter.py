"""F-08: 업체 목록 엑셀 다운로드"""
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


HEADERS = ["업체명", "업종", "홈페이지 URL", "이메일", "검증 상태", "URL 상태", "카테고리", "순위 구간"]
STATUS_KO = {"confirmed": "확인됨", "estimated": "추정", "unknown": "미확인",
             "accessible": "접속가능", "inaccessible": "불가", "needs_review": "재확인필요"}


def _write_sheet(ws, data: list[dict]):
    # 헤더 스타일
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for c in data:
        ws.append([
            c.get("company_name", ""),
            c.get("industry", ""),
            c.get("website_url", ""),
            c.get("email", ""),
            STATUS_KO.get(c.get("email_status", ""), c.get("email_status", "")),
            STATUS_KO.get(c.get("url_status", ""), c.get("url_status", "")),
            c.get("category", ""),
            c.get("rank_range", ""),
        ])

    # 컬럼 너비 자동 조정
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


def export_xlsx(companies: list[dict], sheet_mode: str = "all") -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    sheets = {
        "all": ("전체", companies),
        "confirmed": ("검증완료", [c for c in companies if c.get("email_status") == "confirmed"]),
        "unknown": ("미확인", [c for c in companies if c.get("email_status") in ("unknown", None)]),
    }

    for mode, (name, data) in sheets.items():
        if sheet_mode not in ("all", mode):
            continue
        ws = wb.create_sheet(name)
        _write_sheet(ws, data)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def get_filename() -> str:
    return f"coldmail_list_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
